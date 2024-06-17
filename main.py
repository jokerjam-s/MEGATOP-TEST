import json

import aiohttp
import asyncio

import openpyxl
from openpyxl.styles import Border, Side, Font, PatternFill

from schemas import Category, Product, ProductsInfo

BASE_URL = 'https://www.wildberries.ru/'
CATALOG_URL = 'https://static-basket-01.wbbasket.ru/vol0/data/main-menu-ru-ru-v2.json'

MAX_LEVEL_NESTING = 99

REPORT_FILE_NAME = 'data.xlsx'


async def parse_category(
        data: dict,
        level: int,
        max_level_nesting: int = MAX_LEVEL_NESTING
) -> Category | None:
    if not data or not data.get('isDenyLink') or level > max_level_nesting:
        return None

    try:
        parse_object = Category(**data)
    except Exception as e:
        print(f'Error: {e}')
        return None

    return parse_object


async def parse_products(category: Category, prods_list_result: list[ProductsInfo], category_path: str = ''):
    if category.childs:
        for child in category.childs:
            await parse_products(child, prods_list_result, category_path + f'{category.name} - ')
    else:
        products_info = ProductsInfo()
        products_info.categories_full_path = category_path + category.name
        page_no = 1
        params = {
            'cat': category.id,
            'limit': 100,
            'sort': 'popular',
            'page': page_no,
            'appType': 128,
            'curr': 'byn',
            'lang': 'ru',
            'dest': -59208,
            'spp': 30,
        }

        url = f'https://catalog.wb.ru/catalog/{category.shard}/v1/catalog'

        async with aiohttp.ClientSession() as session:
            async with session.get(url, params=params) as response:
                if response.status == 200:
                    try:
                        data = json.loads(await response.text())
                        products = data.get('data', {}).get('products')
                        products_info.prods_list.extend([Product(**p) for p in products])
                    except Exception as e:
                        print(f'Error: {e}')

        prods_list_result.append(products_info)


async def save_to_excel(products: list[ProductsInfo], file_name: str) -> None:
    work_book = openpyxl.Workbook(iso_dates=True)
    work_book.remove(work_book.worksheets[0])

    thins = Side(border_style="thin", color="000000")
    border_cells = Border(left=thins, right=thins, top=thins, bottom=thins)
    row_no = 1

    for prod_info in products:
        try:
            work_sheet = work_book[prod_info.categories_full_path.split(' - ')[0]]
        except KeyError:
            row_no = 1
            col_no = 1
            work_sheet = work_book.create_sheet(prod_info.categories_full_path.split(' - ')[0])
            for fl in Product.__fields__:
                work_sheet.cell(row=row_no, column=col_no, value=fl).border = border_cells
                work_sheet.cell(row=row_no, column=col_no).fill = PatternFill("solid", fgColor="ABDB77")
                col_no += 1
            row_no += 1
        except Exception as e:
            print(f'Error: {e}')
            return

        col_no = 1
        work_sheet.cell(row=row_no, column=col_no).value = prod_info.categories_full_path
        work_sheet.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=len(Product.__fields__))
        work_sheet.cell(row=row_no, column=col_no).border = border_cells
        work_sheet.cell(row=row_no, column=col_no).font = Font(bold=True)
        work_sheet.cell(row=row_no, column=col_no).fill = PatternFill("solid", fgColor="B6EC66")
        row_no += 1

        for prod in prod_info.prods_list:
            work_sheet.cell(row=row_no, column=1, value=prod.id).border = border_cells
            work_sheet.cell(row=row_no, column=2, value=prod.dist).border = border_cells
            work_sheet.cell(row=row_no, column=3, value=prod.brand).border = border_cells
            if prod.colors:
                work_sheet.cell(row=row_no, column=4).value = ', '.join([str(c) for c in prod.colors])
            work_sheet.cell(row=row_no, column=4).border = border_cells
            work_sheet.cell(row=row_no, column=5, value=prod.name).border = border_cells
            work_sheet.cell(row=row_no, column=6, value=prod.supplier).border = border_cells
            work_sheet.cell(row=row_no, column=7, value=prod.supplierId).border = border_cells
            work_sheet.cell(row=row_no, column=8, value=prod.supplierRating).border = border_cells
            work_sheet.cell(row=row_no, column=9, value=prod.rating).border = border_cells
            work_sheet.cell(row=row_no, column=10, value=prod.reviewRating).border = border_cells
            work_sheet.cell(row=row_no, column=11, value=prod.volume).border = border_cells
            if prod.sizes:
                work_sheet.cell(row=row_no, column=12).value = ', '.join([str(s) for s in prod.sizes])
            work_sheet.cell(row=row_no, column=12).border = border_cells
            work_sheet.cell(row=row_no, column=13, value=prod.priceU).border = border_cells
            work_sheet.cell(row=row_no, column=14, value=prod.totalQuantity).border = border_cells
            row_no += 1

        for letter in range(ord('A'), ord('A') + len(Product.__fields__)):
            work_sheet.column_dimensions[chr(letter)].auto_size = True

    work_book.save(file_name)


async def main():
    categories_data = None
    async with aiohttp.ClientSession() as session:
        async with session.get(CATALOG_URL) as response:
            try:
                categories_data = json.loads(await response.text())
            except json.decoder.JSONDecodeError as e:
                print(f'Error: {e}')

    if categories_data:
        categories = [c for c in [await parse_category(d, 0) for d in categories_data] if c is not None]
        prods = []
        for category in categories:
            await parse_products(category, prods_list_result=prods)
        await save_to_excel(prods, REPORT_FILE_NAME)


if __name__ == '__main__':
    asyncio.run(main())
